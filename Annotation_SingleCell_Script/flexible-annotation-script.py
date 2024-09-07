import sys
import pandas as pd
import os
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def read_input_file(file_path):
    """Read input file based on its extension."""
    print(f"Reading input file: {file_path}")
    _, ext = os.path.splitext(file_path)
    if ext.lower() == '.csv':
        return pd.read_csv(file_path, encoding='unicode_escape')
    elif ext.lower() == '.txt':
        return pd.read_csv(file_path, sep="\t", encoding='unicode_escape')
    else:
        raise ValueError(f"Unsupported file format: {ext}. Please use .csv or .txt files.")

def convert_gene_names(gene_series, species):
    """Convert gene names based on species."""
    print(f"Converting gene names for species: {species}")
    if species.lower() == 'mouse':
        return gene_series.str.capitalize()
    elif species.lower() == 'human':
        return gene_series.str.upper()
    else:
        raise ValueError("Invalid species. Please specify 'mouse' or 'human'.")

def format_excel_summary(ws, summary_data, cluster_data):
    """Format the Excel summary sheet."""
    print("Formatting summary sheet")
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Overview section
    ws['A1'] = "Annotation Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    ws['A3'] = "Overview"
    ws['A3'].font = Font(bold=True, size=12)
    ws.merge_cells('A3:C3')

    for i, (key, value) in enumerate(summary_data.items(), start=4):
        ws[f'A{i}'] = key
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # Cluster-wise Information
    start_row = i + 2
    ws[f'A{start_row}'] = "Cluster-wise Information"
    ws[f'A{start_row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{start_row}:F{start_row}')

    headers = ["Cluster", "Total Genes", "Annotated Genes", "Annotation Rate", "Top 5 Genes by avg_log2FC"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row + 2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for i, cluster_info in enumerate(cluster_data, start=start_row + 3):
        for j, value in enumerate(cluster_info, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if j == 5:  # Top 5 genes column
                cell.alignment = Alignment(wrapText=True)

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 15
    ws.column_dimensions['E'].width = 40

    print("Summary formatting complete")

# Get input and output file paths from command line arguments
inmark = sys.argv[1]
outmark = sys.argv[2]
lmarkers = sys.argv[3]
species = sys.argv[4]  # Species argument

print(f"Processing annotation for {species} genes")

# Read input files
markers1 = read_input_file(lmarkers)
infile = read_input_file(inmark)

# Use the correct column names for markers file
gene_column_markers = 'Genes'
subtype_column_markers = 'Subtype'

# Convert gene names and save markers
print("Converting gene names in marker file")
markers1[gene_column_markers] = convert_gene_names(markers1[gene_column_markers], species)

# Create dictionary from markers
print("Creating marker dictionary")
d = {}
for _, row in markers1.iterrows():
    gene = row[gene_column_markers]
    subtype = row[subtype_column_markers]
    d.setdefault(gene, []).append(subtype)

e = {i: str(j) for i, j in d.items()}
markers = pd.Series(e).to_frame(subtype_column_markers)
markers.reset_index(inplace=True)
markers = markers.rename(columns={'index': gene_column_markers})

# Use the correct column name for gene in the input file
gene_column_infile = 'gene'

# Convert gene names in the input file
print("Converting gene names in input file")
infile[gene_column_infile] = convert_gene_names(infile[gene_column_infile], species)

# Get unique cluster identifiers
unique_clusters = infile['cluster'].unique()
print(f"Found {len(unique_clusters)} unique clusters")

# Initialize counters and lists for summary
total_genes = 0
total_annotated_genes = 0
cluster_gene_counts = {}
cluster_annotated_counts = {}
top_genes_per_cluster = {}

# Process each cluster and save to the Excel file
print("Processing clusters and saving to Excel")
with pd.ExcelWriter(outmark, engine="openpyxl") as writer:
    for cluster in unique_clusters:
        print(f"Processing cluster {cluster}")
        a = infile[infile["cluster"] == cluster]
        b = a.merge(markers, left_on=gene_column_infile, right_on=gene_column_markers, how="left")
        
        # Count genes and annotated genes
        cluster_gene_counts[cluster] = len(b)
        cluster_annotated_counts[cluster] = b[subtype_column_markers].notna().sum()
        total_genes += cluster_gene_counts[cluster]
        total_annotated_genes += cluster_annotated_counts[cluster]
        
        # Get top 5 genes by avg_log2FC
        if 'avg_log2FC' in b.columns:
            top_genes = b.nlargest(5, 'avg_log2FC')[gene_column_infile].tolist()
        else:
            top_genes = ['avg_log2FC column not found']
        top_genes_per_cluster[cluster] = top_genes
        
        b.to_excel(writer, sheet_name=f"cluster_{cluster}", index=False)

    # Generate summary information
    print("Generating summary information")
    summary_data = {
        "Species": species.capitalize(),
        "Total number of genes processed": total_genes,
        "Total number of annotated genes": total_annotated_genes,
        "Overall annotation rate": f"{(total_annotated_genes/total_genes)*100:.2f}%"
    }

    cluster_data = []
    for cluster in unique_clusters:
        cluster_data.append([
            cluster,
            cluster_gene_counts[cluster],
            cluster_annotated_counts[cluster],
            f"{(cluster_annotated_counts[cluster]/cluster_gene_counts[cluster])*100:.2f}%",
            ", ".join(top_genes_per_cluster[cluster])
        ])

    # Create and format summary sheet
    print("Creating and formatting summary sheet")
    writer.book.create_sheet("Summary", 0)
    summary_sheet = writer.book["Summary"]
    format_excel_summary(summary_sheet, summary_data, cluster_data)

print(f"Annotation complete. Results and summary saved to: {outmark}")

# Command to run the script
# python /path/to/annotate.py /path/to/input_file.csv /path/to/output_file.xlsx /path/to/marker_genes.txt species
