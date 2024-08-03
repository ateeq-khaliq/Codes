import os
from openpyxl import load_workbook, Workbook

# Path to the original Excel file
input_file = "/Users/akhaliq/Desktop/new_manuscript/Editors_final/Editors_check_V2/Suppli_tables/Supplementary_TablesV8.xlsx"

# Directory where the split files will be saved
output_dir = os.path.dirname(input_file)

print(f"Input file: {input_file}")
print(f"Output directory: {output_dir}")

# Check if the input file exists
if not os.path.exists(input_file):
    print(f"Error: Input file does not exist: {input_file}")
    exit(1)

try:
    # Load the workbook
    print("Loading workbook...")
    wb = load_workbook(input_file)
    print(f"Workbook loaded. Sheets found: {wb.sheetnames}")

    # Iterate through each sheet
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        
        # Create a new workbook
        new_wb = Workbook()
        # Remove the default sheet created with the new workbook
        new_wb.remove(new_wb.active)
        # Copy the sheet to the new workbook
        new_wb.create_sheet(title=sheet_name)
        new_sheet = new_wb[sheet_name]
        
        # Copy the contents of the sheet
        for row in wb[sheet_name].iter_rows(values_only=True):
            new_sheet.append(row)
        
        # Save the new workbook with the sheet name
        new_filename = f"{sheet_name}.xlsx"
        new_filepath = os.path.join(output_dir, new_filename)
        
        try:
            new_wb.save(new_filepath)
            print(f"Created: {new_filepath}")
        except PermissionError:
            print(f"Error: Permission denied when trying to save {new_filepath}")
        except Exception as e:
            print(f"Error saving {new_filepath}: {str(e)}")

    print("Splitting complete!")

except Exception as e:
    print(f"An error occurred: {str(e)}")
